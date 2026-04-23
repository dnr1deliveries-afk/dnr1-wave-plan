"""
wave_plan_excel_generator.py — Generates print-optimized WAVE PLAN Excel file.

Produces a space-efficient Excel file matching the Yard Marshal wave plan format:
- Full header structure preserved (Wave title, timing labels, timing values, column headers)
- Empty lane rows removed for print efficiency
- Lane assignments preserved EXACTLY as output by lane_optimizer
- Professional formatting (yellow headers, green data rows)
- Print-ready layout (landscape, fit to page)

IMPORTANT: This generator does NOT re-sort or re-assign lanes.
The lane_optimizer assigns lanes, wave_engine sorts by lane_num,
and this generator preserves that exact order.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
#  STYLING CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Colors
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Fonts
BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=12)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────

def generate_wave_plan_excel(wave_plan, output_path=None, station="DNR1"):
    """
    Generate a print-optimized WAVE PLAN Excel file.
    
    Args:
        wave_plan: Wave plan dict from wave_engine.build_wave_plan()
        output_path: Output file path (default: WavePlan_{station}_{date}.xlsx)
        station: Station code for filename
    
    Returns:
        Path to generated Excel file
    """
    if output_path is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_path = f"WavePlan_{station}_{date_str}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "WAVE PLAN"
    
    # Set column widths
    col_widths = [2, 14, 22, 10, 12, 14, 22, 10, 12]  # A is spacer
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    current_row = 1
    waves = wave_plan.get("waves", [])
    
    for wave_idx, wave in enumerate(waves):
        wave_num = wave.get("wave_number", wave_idx + 1)
        
        # Get pad data
        pad_a = wave.get("pad_a", {})
        pad_b = wave.get("pad_b", {})
        
        pad_a_routes = pad_a.get("routes", [])
        pad_b_routes = pad_b.get("routes", [])
        
        # Write wave block - routes are ALREADY sorted by lane_num from wave_engine
        current_row = _write_wave_block(
            ws, current_row, wave_num,
            pad_a_routes, pad_b_routes,
            pad_a, pad_b
        )
        
        # Add spacing between waves
        current_row += 1
    
    # Set print settings
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # Save
    wb.save(output_path)
    return output_path


def _write_wave_block(ws, start_row, wave_num, pad_a_routes, pad_b_routes, 
                       pad_a_info, pad_b_info):
    """
    Write a single wave block (both pads) to the worksheet.
    
    IMPORTANT: Routes are written in the EXACT order they are passed in.
    The lane_optimizer and wave_engine have already determined the correct
    lane assignments and sort order. We do NOT re-sort here.
    
    Returns the next available row number.
    """
    row = start_row
    
    # ─── Row 1: Wave Title ───
    ws.merge_cells(f'B{row}:E{row}')
    ws.merge_cells(f'F{row}:I{row}')
    
    ws[f'B{row}'] = f"WAVE {wave_num} PAD A"
    ws[f'F{row}'] = f"WAVE {wave_num}                             PAD B"
    
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        cell = ws[f'{col}{row}']
        cell.fill = YELLOW_FILL
        cell.font = TITLE_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    row += 1
    
    # ─── Row 2: Timing Labels ───
    timing_labels = ["FIRST ENTRANCE:", "LAST ENTRANCE:", "WAVE TIME:", "LAST EXIT:"]
    for i, label in enumerate(timing_labels):
        ws.cell(row=row, column=2+i, value=label).fill = YELLOW_FILL
        ws.cell(row=row, column=2+i).font = BOLD_FONT
        ws.cell(row=row, column=2+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2+i).border = THIN_BORDER
        
        ws.cell(row=row, column=6+i, value=label).fill = YELLOW_FILL
        ws.cell(row=row, column=6+i).font = BOLD_FONT
        ws.cell(row=row, column=6+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=6+i).border = THIN_BORDER
    
    row += 1
    
    # ─── Row 3: Timing Values ───
    pad_a_values = [
        pad_a_info.get("first_entrance", ""),
        pad_a_info.get("last_entrance", ""),
        pad_a_info.get("wave_time", ""),
        pad_a_info.get("last_exit", "")
    ]
    pad_b_values = [
        pad_b_info.get("first_entrance", ""),
        pad_b_info.get("last_entrance", ""),
        pad_b_info.get("wave_time", ""),
        pad_b_info.get("last_exit", "")
    ]
    
    for i, val in enumerate(pad_a_values):
        ws.cell(row=row, column=2+i, value=val).fill = YELLOW_FILL
        ws.cell(row=row, column=2+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2+i).border = THIN_BORDER
    
    for i, val in enumerate(pad_b_values):
        ws.cell(row=row, column=6+i, value=val).fill = YELLOW_FILL
        ws.cell(row=row, column=6+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=6+i).border = THIN_BORDER
    
    row += 1
    
    # ─── Row 4: Column Headers ───
    col_headers = ["Staging Lane", "Route No.", "No. Carts", "Notes"]
    for i, header in enumerate(col_headers):
        ws.cell(row=row, column=2+i, value=header).fill = YELLOW_FILL
        ws.cell(row=row, column=2+i).font = BOLD_FONT
        ws.cell(row=row, column=2+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2+i).border = THIN_BORDER
        
        ws.cell(row=row, column=6+i, value=header).fill = YELLOW_FILL
        ws.cell(row=row, column=6+i).font = BOLD_FONT
        ws.cell(row=row, column=6+i).alignment = CENTER_ALIGN
        ws.cell(row=row, column=6+i).border = THIN_BORDER
    
    row += 1
    
    # ─── Data Rows ───
    # Filter out cargo bike routes (BK_ prefix) - they go in STG-C section
    # DO NOT RE-SORT - preserve exact order from lane_optimizer/wave_engine
    pad_a_van = [r for r in pad_a_routes if r.get("route") and not r.get("route", "").startswith("BK_")]
    pad_b_van = [r for r in pad_b_routes if r.get("route") and not r.get("route", "").startswith("BK_")]
    
    # Calculate rows needed (only for routes that exist)
    max_data_rows = max(len(pad_a_van), len(pad_b_van))
    
    # Write data rows in EXACT order (preserving lane_optimizer assignments)
    for i in range(max_data_rows):
        is_alt_row = (i % 2 == 0)
        row_fill = GREEN_FILL if is_alt_row else LIGHT_GREEN_FILL
        
        # Pad A
        if i < len(pad_a_van):
            route = pad_a_van[i]
            _write_route_row(ws, row, 2, route, row_fill)
        else:
            _write_empty_row(ws, row, 2, row_fill)
        
        # Pad B
        if i < len(pad_b_van):
            route = pad_b_van[i]
            _write_route_row(ws, row, 6, route, row_fill)
        else:
            _write_empty_row(ws, row, 6, row_fill)
        
        row += 1
    
    # ─── STG-C Rows (Cargo Bikes) ───
    # Only include if there are cargo bike routes in this wave
    pad_a_cargo = [r for r in pad_a_routes if r.get("route", "").startswith("BK_")]
    pad_b_cargo = [r for r in pad_b_routes if r.get("route", "").startswith("BK_")]
    
    if pad_a_cargo or pad_b_cargo:
        max_cargo = max(len(pad_a_cargo), len(pad_b_cargo), 5)  # At least 5 STG-C rows
        for i in range(max_cargo):
            # Pad A
            if i < len(pad_a_cargo):
                route = pad_a_cargo[i]
                if not route.get("lane"):
                    route["lane"] = f"STG-C{i+1}"
                _write_route_row(ws, row, 2, route, WHITE_FILL)
            else:
                ws.cell(row=row, column=2, value=f"STG-C{i+1}").border = THIN_BORDER
                for col in [3, 4, 5]:
                    ws.cell(row=row, column=col).border = THIN_BORDER
            
            # Pad B
            if i < len(pad_b_cargo):
                route = pad_b_cargo[i]
                if not route.get("lane"):
                    route["lane"] = f"STG-C{i+1}"
                _write_route_row(ws, row, 6, route, WHITE_FILL)
            else:
                ws.cell(row=row, column=6, value=f"STG-C{i+1}").border = THIN_BORDER
                for col in [7, 8, 9]:
                    ws.cell(row=row, column=col).border = THIN_BORDER
            
            row += 1
    
    # ─── Total Row ───
    pad_a_total = len(pad_a_van)
    pad_b_total = len(pad_b_van)
    
    ws.cell(row=row, column=2, value="TOTAL ROUTES:").font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=4, value=pad_a_total).font = BOLD_FONT
    ws.cell(row=row, column=4).alignment = CENTER_ALIGN
    ws.cell(row=row, column=4).border = THIN_BORDER
    ws.cell(row=row, column=5).border = THIN_BORDER
    
    ws.cell(row=row, column=6, value="TOTAL ROUTES:").font = BOLD_FONT
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=7).border = THIN_BORDER
    ws.cell(row=row, column=8, value=pad_b_total).font = BOLD_FONT
    ws.cell(row=row, column=8).alignment = CENTER_ALIGN
    ws.cell(row=row, column=8).border = THIN_BORDER
    ws.cell(row=row, column=9).border = THIN_BORDER
    
    row += 1
    
    return row


def _write_route_row(ws, row, start_col, route, fill):
    """Write a single route data row."""
    lane = route.get("lane", route.get("lane_label", ""))
    route_no = _format_route_display(route)
    carts = route.get("total_carts", 0)
    notes = _get_notes(route)
    
    # Staging Lane
    cell = ws.cell(row=row, column=start_col, value=lane)
    cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN
    
    # Route No.
    cell = ws.cell(row=row, column=start_col+1, value=route_no)
    cell.fill = fill
    cell.border = THIN_BORDER
    
    # No. Carts
    cell = ws.cell(row=row, column=start_col+2, value=carts if carts else "")
    cell.fill = fill
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN
    
    # Notes
    cell = ws.cell(row=row, column=start_col+3, value=notes)
    if notes == "Large Van":
        cell.fill = ORANGE_FILL
    else:
        cell.fill = fill
    cell.border = THIN_BORDER


def _write_empty_row(ws, row, start_col, fill):
    """Write an empty row with borders."""
    for i in range(4):
        cell = ws.cell(row=row, column=start_col+i)
        cell.fill = fill
        cell.border = THIN_BORDER


def _format_route_display(route):
    """Format route for display: 'DSP - CA_A155 (S)'"""
    dsp = route.get("dsp", "")
    route_code = route.get("route", "")
    
    # Get service type abbreviation
    abbrev = route.get("service_abbrev", "")
    if not abbrev:
        service_type = route.get("service_type", "")
        abbrev = _get_service_abbrev(service_type)
    
    if dsp and route_code:
        return f"{dsp} - {route_code} ({abbrev})"
    elif route_code:
        return f"{route_code} ({abbrev})"
    return ""


def _get_service_abbrev(service_type):
    """Get single-letter service type abbreviation."""
    if not service_type:
        return "S"
    
    st_lower = service_type.lower()
    if "large van" in st_lower:
        return "L"
    elif "low emission" in st_lower or "lev" in st_lower:
        return "LEV"
    elif "nursery" in st_lower:
        return "N"
    elif "remote" in st_lower or "debrief" in st_lower:
        return "R"
    elif "flex" in st_lower:
        return "F"
    else:
        return "S"  # Standard


def _get_notes(route):
    """Get notes for route (e.g., 'Large Van')."""
    service_type = route.get("service_type", "")
    service_abbrev = route.get("service_abbrev", "")
    
    if service_abbrev == "L" or (service_type and "large van" in service_type.lower()):
        return "Large Van"
    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  STANDALONE TEST
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Test with sample data matching wave_engine output structure
    # Routes are pre-sorted by lane_num (as wave_engine does)
    sample_wave_plan = {
        "waves": [
            {
                "wave_number": 1,
                "wave_label": "Wave 1",
                "pad_a": {
                    "pad_label": "A",
                    "wave_time": "10:20",
                    "first_entrance": "10:05",
                    "last_entrance": "10:15",
                    "last_exit": "10:40",
                    "routes": [
                        # Already sorted by lane_num ascending (wave 1)
                        {"route": "CA_A155", "dsp": "MOLI", "lane": "STG-A1", "lane_num": 1, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                        {"route": "CA_A154", "dsp": "MOLI", "lane": "STG-A2", "lane_num": 2, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                        {"route": "CA_A152", "dsp": "MOLI", "lane": "STG-A3", "lane_num": 3, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                    ],
                    "total_routes": 3,
                    "total_carts": 6,
                },
                "pad_b": {
                    "pad_label": "B",
                    "wave_time": "10:30",
                    "first_entrance": "10:15",
                    "last_entrance": "10:25",
                    "last_exit": "10:50",
                    "routes": [
                        {"route": "CA_A305", "dsp": "HPLM", "lane": "STG-B1", "lane_num": 1, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                        {"route": "CA_A307", "dsp": "HPLM", "lane": "STG-B2", "lane_num": 2, "total_carts": 3, "service_type": "Standard Parcel - Large Van", "service_abbrev": "L"},
                    ],
                    "total_routes": 2,
                    "total_carts": 5,
                },
                "status": "not_started",
            },
            {
                "wave_number": 2,
                "wave_label": "Wave 2",
                "pad_a": {
                    "pad_label": "A",
                    "wave_time": "10:45",
                    "first_entrance": "10:30",
                    "last_entrance": "10:40",
                    "last_exit": "11:05",
                    "routes": [
                        # Already sorted by lane_num ascending (wave_engine does this)
                        # The physical sheet may display differently but data is by lane_num
                        {"route": "CA_A166", "dsp": "MOLI", "lane": "STG-A28", "lane_num": 28, "total_carts": 3, "service_type": "Standard Parcel - Large Van", "service_abbrev": "L"},
                        {"route": "CA_A157", "dsp": "MOLI", "lane": "STG-A29", "lane_num": 29, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                        {"route": "CA_A178", "dsp": "MOLI", "lane": "STG-A30", "lane_num": 30, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                    ],
                    "total_routes": 3,
                    "total_carts": 7,
                },
                "pad_b": {
                    "pad_label": "B",
                    "wave_time": "10:55",
                    "first_entrance": "10:40",
                    "last_entrance": "10:50",
                    "last_exit": "11:15",
                    "routes": [
                        {"route": "CA_A300", "dsp": "HPLM", "lane": "STG-B29", "lane_num": 29, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                        {"route": "CA_A306", "dsp": "HPLM", "lane": "STG-B30", "lane_num": 30, "total_carts": 2, "service_type": "Standard Parcel Medium Van", "service_abbrev": "S"},
                    ],
                    "total_routes": 2,
                    "total_carts": 4,
                },
                "status": "not_started",
            }
        ],
        "wave_c": [],
        "summary": {
            "total_waves": 2,
            "total_routes": 10,
            "total_carts": 22,
        }
    }
    
    output = generate_wave_plan_excel(sample_wave_plan, "test_wave_plan.xlsx")
    print(f"Generated: {output}")
